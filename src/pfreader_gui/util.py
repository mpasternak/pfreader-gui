import sys
from io import BytesIO

import openpyxl
from PyQt5.QtCore import QSortFilterProxyModel
from pfreader.core import dir_contains_pflex_data, get_year_dirs


def autofit_databook(db):
    """Autofit databook columns."""
    wb_in = openpyxl.load_workbook(BytesIO(db.xlsx))
    wb_out = openpyxl.Workbook()
    for idx, sheet_name in enumerate(wb_in.sheetnames):

        ws_in = wb_in.get_sheet_by_name(sheet_name)
        ws_out = wb_out.create_sheet(sheet_name, idx)
        ws_out.freeze_panes = 'A2'

        for row in ws_in.rows:
            new_row = []
            for value in row:
                new_row.append(value.value)
            ws_out.append(new_row)

        for column_cells in ws_out.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws_out.column_dimensions[column_cells[0].column].width = length

    return wb_out


class ExcludeSomeNamesModel(QSortFilterProxyModel):
    def filterAcceptsRow(self, p_int, idx):
        index = self.sourceModel().index(p_int, 0, idx)
        s = self.sourceModel().data(index)

        if sys.platform == 'darwin':
            if s.startswith("Preboot"):
                return False

            if s.startswith("Volumes"):
                return True

        path = self.sourceModel().filePath(index)

        contains = dir_contains_pflex_data(path)
        if contains:
            return True

        older_prisma = get_year_dirs(path)
        for elem in older_prisma:
            # There may be year dirs in the root directory for older machines
            return True

        return False

    pass
